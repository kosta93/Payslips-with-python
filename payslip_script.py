#library to import the excel file
import openpyxl
#libraries to create the pdf file and add text to it
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfbase.ttfonts import TTFont
#libraries to merge pdf files
import os
from PyPDF2 import PdfFileReader, PdfFileMerger

#convert the font so it is compatible
pdfmetrics.registerFont(TTFont('Arial','Arial.ttf'))

#import the sheet from the excel file
wb = openpyxl.load_workbook('C:\\Users\\Gebruiker\\Desktop\\tutorial\\data.xlsx')
sheet = wb.get_sheet_by_name('employees')

#Page information
page_width = 2156
page_height = 3050
spread = 100
start = 200
start_2 = 700

#Payslip variables
company_name = 'The best company in the worldW'
month_year = 'August 2019'

def create_payslip():
    for i in range (2, 42):
        #eading values from excel file
        emp_id = sheet.cell(row = i, column = 1).value
        emp_name = sheet.cell(row = i, column = 2).value
        emp_last_name = sheet.cell(row = i, column = 3).value
        gross_salary = sheet.cell(row = i, column = 4).value
        pension_contr = sheet.cell(row = i, column = 5).value
        health_ins = sheet.cell(row = i, column = 6).value
        p_i_t = sheet.cell(row = i, column = 7).value
        bonus_payment = sheet.cell(row = i, column = 8).value
        deduction = sheet.cell(row = i, column = 9).value
        net_salary = sheet.cell(row = i, column = 10).value

        #Creating a pdf file and setting a naming convention
        c = canvas.Canvas(str(emp_name)+'_'+str(emp_last_name)+'_'+str(emp_id) + '_' + month_year + '.pdf' )
        #Page settings (size/font)
        c.setPageSize((page_width, page_height))
        c.setFont('Arial',80)
        #Company name text
        text_width = stringWidth(company_name, 'Arial',80)
        c.drawString((page_width-text_width)/2, 2900, company_name)
        #Invoice month/year information
        text = 'Salary calculation for period ' + month_year
        text_width = stringWidth(text, 'Arial',55)
        c.setFont('Arial',55)
        c.drawString((page_width-text_width)/2, 2700, text)

        y = 2500
        #Drawing payslip related information
        c.setFont('Arial',45)
        c.drawString(start, y, 'Employee\'s id: ')
        c.drawString(start_2, y, str(emp_id))
        y -= spread

        c.drawString(start, y, 'Employee\'s name:')
        c.drawString(start_2, y, str(emp_name) + ' ' + str(emp_last_name))
        y -= spread

        c.drawString(start, y, 'Gross salary:')
        c.drawString(start_2, y, str(gross_salary))
        y -= spread

        c.drawString(start, y, 'Pension contribution:')
        c.drawString(start_2, y, str(pension_contr))
        y -= spread

        c.drawString(start, y, 'Health insurance:')
        c.drawString(start_2, y, str(health_ins))
        y -= spread

        c.drawString(start, y, 'Personal income tax:')
        c.drawString(start_2, y, str(p_i_t))
        y -= spread
        
        c.drawString(start, y, 'Bonus payment:')
        c.drawString(start_2, y, str(bonus_payment))
        y -= spread

        c.drawString(start, y, 'Deduction:')
        c.drawString(start_2, y, str(deduction))
        y -= spread

        c.drawString(start, y, 'Net salary:')
        c.drawString(start_2, y, str(net_salary))
        y -= spread * 3

        c.drawString(start, y, 'Signature: ')
        c.drawString(start_2, y,'____________')
      
        #Saving the pdf file
        c.save()

def merge_pdfs():
    files_dir = 'C:\\Users\\Gebruiker\\Desktop\\tutorial' #Select the directory where the pdf files are located
    pdf_files = [f for f in os.listdir(files_dir) if f.endswith('.pdf')] #Get all files in the directory that end with '.pdf'
    merger = PdfFileMerger() #Create an empty file
    for filename in pdf_files:
        merger.append(PdfFileReader(os.path.join(files_dir,filename),'rb')) #Add every pdf to the empty file
    merger.write(os.path.join(files_dir,'merged_pdfs.pdf')) #Save the file
    

create_payslip()
merge_pdfs()
